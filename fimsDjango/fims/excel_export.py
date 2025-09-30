import io
import os
import datetime
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.shortcuts import get_object_or_404
from .models import FamilyHead, FamilyMember

def excel_view(request):
    head_id = request.GET.get('id')
    if head_id:
        from .utils import decode_id
        head_id = decode_id(head_id)
    if not head_id:
        # Return an empty Excel file with a message
        wb = Workbook()
        ws = wb.active
        ws.title = "Family Report"
        ws.append(["No family selected. Please provide a valid family head ID."])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return FileResponse(output, as_attachment=True, filename="family_report.xlsx")

    head = get_object_or_404(FamilyHead, HeadID=head_id)
    members = FamilyMember.objects.filter(HeadID=head)
    wb = Workbook()
    ws = wb.active
    ws.title = "Family Report"

    # --- Head Table ---
    # Get state and city names if possible
    from .models import State, City
    state_name = head.State
    city_name = head.City
    # Try to resolve state and city names if they are IDs or codes
    try:
        state_obj = State.objects.filter(name=head.State).first() or State.objects.filter(id=head.State).first()
        if state_obj:
            state_name = state_obj.name
    except Exception:
        pass
    try:
        city_obj = City.objects.filter(name=head.City).first() or City.objects.filter(id=head.City).first()
        if city_obj:
            city_name = city_obj.name
    except Exception:
        pass

    head_table = [
        ["Name", f"{head.Name or '-'} {head.Surname or '-'}"],
        ["Birthdate", str(head.Birthdate) if head.Birthdate else '-'],
        ["Gender", head.Gender or '-'],
        ["Mobile", head.MobileNo or '-'],
        ["Address", head.Address or '-'],
        ["State", state_name or '-'],
        ["City", city_name or '-'],
        ["Pincode", head.Pincode or '-'],
        ["Education", head.Education or '-'],
        ["Marital Status", head.MaritalStatus or '-'],
        ["Wedding Date", str(head.WeddingDate) if head.WeddingDate else 'Not Married'],
        ["Hobbies", ', '.join([h.Hobby for h in head.head_hobbies.all()]) if head.head_hobbies.exists() else 'None'],
    ]
    ws.append(["Family Head", ""])
    for row in head_table:
        ws.append(row)
    ws.append([""])

    # --- Family Members Table ---
    ws.append(["Family Members"])
    member_headers = ["Name", "Birthdate", "Gender", "Mobile", "Education", "Marital Status", "Wedding Status"]
    ws.append(member_headers)
    for member in members:
        wedding_status = 'Married' if getattr(member, 'MaritalStatus', None) == 'Married' else 'Not Married'
        ws.append([
            f"{member.Name or '-'} {member.Surname or '-'}",
            str(member.Birthdate) if member.Birthdate else '-',
            member.Gender or '-',
            member.MobileNo or '-',
            '-',  # No Education field for FamilyMember
            member.MaritalStatus or '-',
            wedding_status,
        ])

    # --- Styling ---
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center
            cell.border = border
    for cell in ws[2]:
        cell.font = bold
    for cell in ws[ws.max_row]:
        cell.font = bold
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Family Members":
            for cell in row:
                cell.font = bold
                cell.fill = fill
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"family_report_{head.Name}_{head.Surname}_{head_id}.xlsx"
    return FileResponse(output, as_attachment=True, filename=filename)
